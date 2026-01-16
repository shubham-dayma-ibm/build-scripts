import os
import shutil
import subprocess
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from config import *
from openpyxl import load_workbook
import ibm_boto3
from ibm_botocore.client import Config
from pathlib import Path
from package_tags import package_tags
from copy import deepcopy
import re
# =========================
# CONFIGURATION
# =========================

EXCEL_FILE = "packages_to_process_non_noarch-v1.xlsx"
WORK_BASE_DIR = Path.cwd()
MAX_WORKERS = 64  # safe default for 144-core system
REPO_MAIN_DIR = "/root/shubham_playground/"

READ_BUILDINFO_SCRIPT = "read_buildinfo.sh"
BUILD_WHEEL_SCRIPT = "build_wheels.sh"

FILES_TO_COPY = [
        READ_BUILDINFO_SCRIPT,
        BUILD_WHEEL_SCRIPT,
        "create_wheel_wrapper.sh",
        "build_wheels.py",
        "license_suffix.py"
    ]


excel_lock = threading.Lock()

# =========================
# HELPERS
# =========================
global _cos_client

def create_cos_client():
    global _cos_client

    '''Create and return an IBM COS client'''
    _cos_client=ibm_boto3.client(
        "s3",
        ibm_api_key_id=COS_API_KEY,
        ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
        config=Config(signature_version="oauth"),
        endpoint_url=COS_ENDPOINT,
    )

def upload_to_ibm_cos(package, version, wheel_name, source, sha256):
    """
    Placeholder for IBM COS upload
    """
    global _cos_client
    
    desitination = f"{package}/{version}/{wheel_name}"

    # Implement actual upload logic later
    try:
        _cos_client.upload_file(
                    source,
                    COS_BUCKET,
                    desitination,
                    ExtraArgs={
                        "Metadata": {
                            "sha256": sha256
                        }
                    }
                )
        
        return True
    except Exception as e:
        return False


def normalize_docker_name(name: str) -> str:
    # lowercase
    name = name.lower()

    # replace invalid characters with hyphen
    name = re.sub(r"[^a-z0-9_.-]", "-", name)

    # docker name must start with alphanumeric
    name = re.sub(r"^[^a-z0-9]+", "", name)

    # collapse multiple hyphens
    name = re.sub(r"-{2,}", "-", name)

    return name

def run_cmd(cmd, cwd):
    """
    Run shell command and return True/False
    """
    result = subprocess.run(
        cmd,
        cwd=cwd,
        shell=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE
    )
    
    # TODO: remove below lines
    print('===================================')
    print(cmd)
    print(result.stdout.decode())
    print(result.stderr.decode())
    print('===================================')
    
    return result.returncode == 0

def determine_version_suffix(wheel_name):
    if wheel_name in ppc64le2_suffix_list:
        return "ppc64le2"
    return "ppc64le1"

def determine_version_tag(pkg, ver):
    tags = package_tags.get(pkg, [])
    for tag in tags:
        if tag.lstrip("v") == ver:
            ver = tag
    return ver

# =========================
# CORE WORKFLOW
# =========================
def process_row(row_idx, sheet, headers, testing_wheel_dir, build_log_dir):
    pkg = sheet.cell(row=row_idx, column=headers["package_name"]).value
    ver = sheet.cell(row=row_idx, column=headers["package_version"]).value
    pyver = sheet.cell(row=row_idx, column=headers["python_version"]).value
    wheel_name = sheet.cell(row=row_idx, column=headers["wheel_name"]).value
    status_cell = sheet.cell(row=row_idx, column=headers["status"])
    modified_wheel_name_cell = sheet.cell(row=row_idx, column=headers["modified_wheel_name"])
    modified_wheel_sha256_cell = sheet.cell(row=row_idx, column=headers["modified_wheel_sha256"])
    build_script_trigger_with_version_cell = sheet.cell(row=row_idx, column=headers["build_script_trigger_with_version"])

    pkg = pkg.lower()

    work_dir = WORK_BASE_DIR / wheel_name.replace('.whl', "")

    try:
        work_dir.mkdir(parents=True, exist_ok=True)

        # Copy scripts
        for file_to_copy in FILES_TO_COPY:
            shutil.copy(file_to_copy, work_dir)
        
        # Determine suffix (wheel name unknown yet → safe default)
        version_suffix = determine_version_suffix(wheel_name)

        identified_version_tag = determine_version_tag(pkg, deepcopy(ver))

        build_script_trigger_with_version_cell.value = str(identified_version_tag)

        process_container_name = normalize_docker_name(f"{pkg}_{identified_version_tag}_{pyver}")

        # Create export_rebuild_wheel.sh
        export_script = work_dir / "export_rebuild_wheel.sh"
        with export_script.open("w") as f:
            f.write(f"export PACKAGE_NAME={pkg}\n")
            f.write(f"export VERSION={identified_version_tag}\n")
            f.write(f"export PYTHON_VERSION={pyver}\n")
            f.write(f"export VERSION_SUFFIX={version_suffix}\n")
            f.write(f"export REPO_MAIN_DIR={REPO_MAIN_DIR}\n")
            f.write(f"export PROCESS_CONTAINER_NAME={process_container_name}\n")

        # Make executable
        export_script.chmod(0o755)

        # Step 3: read_buildinfo.sh
        if not run_cmd(f"bash {READ_BUILDINFO_SCRIPT}", work_dir):
            with excel_lock:
                status_cell.value = "Failed with read_build_info.sh"
            return

        # Step 4: build_wheel.sh
        if not run_cmd(f"bash {BUILD_WHEEL_SCRIPT}", work_dir):
            if (work_dir / "build_log").exists():
                shutil.move(work_dir / "build_log", f"{build_log_dir}/{wheel_name.replace('.whl', '')}")
            with excel_lock:
                status_cell.value = "Failed with build_wheel.sh"
            return

        # Step 5: audit checks
        if (work_dir / "audit_wheel_skipped").exists():
            shutil.move(work_dir / "build_log", f"{build_log_dir}/{wheel_name.replace('.whl', '')}")
            with excel_lock:
                status_cell.value = "Audit Wheel Skip"
            return

        if (work_dir / "audit_wheel_errored").exists():
            shutil.move(work_dir / "build_log", f"{build_log_dir}/{wheel_name.replace('.whl', '')}")
            with excel_lock:
                status_cell.value = "Audit wheel failed"
            return

        # Step 6: wheel existence
        wheels = list(work_dir.glob("wheelhouse/*.whl"))
        if not wheels:
            shutil.move(work_dir / "build_log", f"{build_log_dir}/{wheel_name.replace('.whl', '')}")
            with excel_lock:
                status_cell.value = "Missing Repaired Wheel"
            return

        # Step 7: upload
        wheel_names = []
        wheel_sha256 = ""
        if (work_dir / "sha256.sha").exists():
            with open(f"{work_dir}/sha256.sha", "r") as f:
                wheel_sha256 = f.read().split()[0]

        for whl in wheels:
            # Let's not upload wheel directly to COS, validate it and later upload it with different script
            # if not upload_to_ibm_cos(pkg, ver, whl.name, whl, sha256):
            #     with excel_lock:
            #         status_cell.value = "Upload Failed"
            #     return

            wheel_names.append(whl.name)
            shutil.move(str(whl), testing_wheel_dir / whl.name)

        with excel_lock:
            status_cell.value = "Success"
            modified_wheel_name_cell.value = ",".join(wheel_names)
            modified_wheel_sha256_cell.value = wheel_sha256
    except:
        import traceback
        traceback.print_exc()
    finally:
        # Cleanup
        shutil.rmtree(work_dir, ignore_errors=True)

# =========================
# MAIN
# =========================
def main():
    # Upload wheels post validation, with different script
    # create_cos_client()
    
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active

    # Create testing wheel dir, if not exist
    testing_wheel_dir = WORK_BASE_DIR / "testing_wheels"
    testing_wheel_dir.mkdir(parents=True, exist_ok=True)

    # Create build_logs dir, if not exist
    build_log_dir = WORK_BASE_DIR / "build_logs"
    build_log_dir.mkdir(parents=True, exist_ok=True)

    # Map headers
    headers = {}
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header:
            headers[header] = col

    # Ensure status column
    if "status" not in headers:
        headers["status"] = sheet.max_column + 1
        sheet.cell(row=1, column=headers["status"]).value = "status"

    if "build_script_trigger_with_version" not in headers:
        headers["build_script_trigger_with_version"] = sheet.max_column + 1
        sheet.cell(row=1, column=headers["build_script_trigger_with_version"]).value = "build_script_trigger_with_version"
    
    # Ensure modified_wheel_name column
    if "modified_wheel_name" not in headers:
        headers["modified_wheel_name"] = sheet.max_column + 1
        sheet.cell(row=1, column=headers["modified_wheel_name"]).value = "modified_wheel_name"
    
    if "modified_wheel_sha256" not in headers:
        headers["modified_wheel_sha256"] = sheet.max_column + 1
        sheet.cell(row=1, column=headers["modified_wheel_sha256"]).value = "modified_wheel_sha256"

    tasks = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        for row_idx in range(2, sheet.max_row + 1):
            status = sheet.cell(row=row_idx, column=headers["status"]).value
            if status:
                continue  # skip already processed rows

            tasks.append(
                executor.submit(process_row, row_idx, sheet, headers, testing_wheel_dir, build_log_dir)
            )

        for _ in as_completed(tasks):
            pass
    
    wb.save(EXCEL_FILE)
    print("✅ Processing completed")


if __name__ == "__main__":
    main()
